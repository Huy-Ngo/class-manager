from openpyxl import load_workbook
from json import dump
from argparse import ArgumentParser
from csv import DictReader

# Taking arguments
parser = ArgumentParser(description='Source data file')
parser.add_argument('--file-name', type=str)

args = parser.parse_args()
file_name = args.file_name
file_type = file_name.split('.')[-1]

# CONSTANTS
CLASS = 'ICT'  # Set to * for all classes

# Universal variable
students = []


def convert_csv():
	with open(file_name, newline='') as f:
		reader = DictReader(f)
		for row in reader:
			students.append({
				'id': row['id'],
				'surname': row['name'],
				'name': row['name'],
				'student-id': row['student-id'],
				'class': row['class'],
			})


# Working with xlsx
def convert_xlsx():
	wb = load_workbook(filename='prob-list.xlsx')
	sheet = wb['students list']

	# Detecting values
	HEADER_COLS = {
		"id": '',
		"surname": '',
		"name": '',
		"full-name": '',
		"student-id": '',
		"class": ''
	}
	header_format = '{}1'

	i = 'A'
	info = sheet[header_format.format(i)].value
	while info != '':
		info = sheet[header_format.format(i)].value
		for j in HEADER_COLS:
			if info == j:
				HEADER_COLS[j] = i
				break
		if i == 'Z':
			break
		ch = ord(i[0])
		ch += 1
		i = chr(ch)

	print(HEADER_COLS)

	last_row = sheet.max_row
	# print(sheet[f'{HEADER_COLS["id"]}{last_row}'].value)
	while sheet[f'{HEADER_COLS["id"]}{last_row}'].value is None:
		last_row -= 1

	for i in range(1, last_row + 1):
		c_id = f'{HEADER_COLS["id"]}{i}'
		c_student_id = f'{HEADER_COLS["student-id"]}{i}'
		c_cls = f'{HEADER_COLS["class"]}{i}'
		c_surname = f'{HEADER_COLS["surname"]}{i}'
		c_name = f'{HEADER_COLS["name"]}{i}'
		c_fname = f'{HEADER_COLS["full-name"]}{i}'

		_id = sheet[c_id].value
		student_id = sheet[c_student_id].value
		cls = sheet[c_cls].value

		if len(HEADER_COLS["full-name"]) == 1:
			# full name, instead of surname + name
			full_name = sheet[c_fname].value
			# print(c_fname, full_name, last_row)
			full_name = full_name.split()
			name = full_name[-1]
			surname = ' '.join(full_name[:-1])
		else:
			surname = sheet[c_surname].value
			name = sheet[c_name].value

		student = {
			"id": _id,
			"surname": surname,
			"name": name,
			"student-id": student_id,
			"class": cls
		}
		if cls == CLASS or CLASS == '*':
			students.append(student)


if file_type == 'xlsx':
	convert_xlsx()
elif file_type == 'csv':
	convert_csv()
else:
	raise Exception(f'Extension *.{file_type} is not supported')

with open('student-list.json', 'w', encoding='utf-8') as f:
	dump(students, f, ensure_ascii=False)
